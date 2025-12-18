#!/usr/bin/env python3
"""
Extract data from Gene's Excel tables to verify dElev values
"""

import openpyxl

def extract_us_he_ring4():
    """Extract US HE Ring 4 data from Excel"""

    wb = openpyxl.load_workbook('Arma Reforger Mortar Calc.xlsx', data_only=True)
    sheet = wb['Range Data']

    print('US M821 HE - Ring 4 Data from Excel (Gene\'s Tables):')
    print('=' * 80)

    # Find the header row
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=50, values_only=False), 1):
        cell_value = row[0].value
        if cell_value and 'US' in str(cell_value) and 'M821' in str(cell_value) and '4 Rings' in str(cell_value):
            print(f'Found header at row {row_idx}: {cell_value}')

            # Get column headers (next row)
            header_row = list(sheet.iter_rows(min_row=row_idx+1, max_row=row_idx+1, values_only=True))[0]
            print(f'Columns: {header_row}')
            print()

            # Get data rows
            headers = ['Range', 'Elev', 'TOF', 'D ELEV']
            print(f"{headers[0]:<8} {headers[1]:<8} {headers[2]:<8} {headers[3]:<12}")
            print('-' * 60)

            data_start = row_idx + 2
            for data_row in sheet.iter_rows(min_row=data_start, max_row=data_start+30, values_only=True):
                if data_row[0] is None:
                    break
                range_val = data_row[0]
                elev_val = data_row[1]
                tof_val = data_row[2]
                delev_val = data_row[3] if len(data_row) > 3 else None

                if range_val:
                    delev_str = str(delev_val) if delev_val else "N/A"
                    print(f'{range_val:<8} {elev_val:<8} {tof_val:<8} {delev_str:<12}')

            break

def extract_marcel_coefficients():
    """Extract polynomial coefficients from Marcel's calculations"""

    print('\n\n')
    print('Marcel\'s Polynomial Coefficients (from second Excel):')
    print('=' * 80)

    try:
        wb = openpyxl.load_workbook('Berechnungen Mor-ohne Map.xlsx', data_only=True)

        print('Available sheets:')
        for sheet_name in wb.sheetnames:
            print(f'  - {sheet_name}')

        print()

        # Look for polynomial data
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f'\nSearching in sheet: {sheet_name}')

            # Search for "Polynome" or "Delta ELEV" keywords
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=100, values_only=True), 1):
                row_str = ' '.join([str(cell) for cell in row if cell])
                if 'Polynome' in row_str or 'Delta ELEV' in row_str or 'Höhe' in row_str:
                    print(f'  Row {row_idx}: {row[:10]}')  # First 10 cells

                    # Print next 5 rows for context
                    for i in range(5):
                        next_row = list(sheet.iter_rows(min_row=row_idx+i+1, max_row=row_idx+i+1, values_only=True))
                        if next_row:
                            print(f'  Row {row_idx+i+1}: {next_row[0][:10]}')

                    break

    except Exception as e:
        print(f'Error reading Marcel\'s file: {e}')

if __name__ == '__main__':
    extract_us_he_ring4()
    extract_marcel_coefficients()
